from __future__ import annotations

import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_FILE = ROOT / "data.json"


def clean_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize_category(value: object) -> str:
    text = clean_text(value)
    return re.sub(r"^\d+\.\s*", "", text) if text else "미분류"


def parse_int(value: object) -> int:
    text = clean_text(value).replace(",", "")
    match = re.search(r"-?\d+", text)
    return int(match.group(0)) if match else 0


def infer_maker(product_code: str, description: str) -> str:
    text = f"{product_code} {description}".lower()
    rules = [
        ("삼성", ["삼성", "samsung", "dm", "db"]),
        ("LG", ["lg", "b70"]),
        ("HP", ["hp", "t640"]),
        ("레노버", ["레노버", "lenovo", "thinkcentre"]),
    ]
    for maker, needles in rules:
        if any(needle.lower() in text for needle in needles):
            return maker
    return ""


def load_existing_state() -> dict:
    if not DATA_FILE.exists():
        return {
            "settings": {
                "warehouseName": "실재고 조사",
                "managerName": "Counter",
                "currencyUnit": "KRW",
                "lowStockDefault": 0,
            },
            "products": [],
            "movements": [],
            "stocktake": {"counts": {}, "updatedAt": ""},
        }
    return json.loads(DATA_FILE.read_text(encoding="utf-8"))


def import_excel(excel_path: Path) -> dict:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    products = []

    for index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        category = normalize_category(row[0] if len(row) > 0 else "")
        product_code = clean_text(row[1] if len(row) > 1 else "")
        description = clean_text(row[2] if len(row) > 2 else "")
        price = clean_text(row[3] if len(row) > 3 else "")
        exposure = clean_text(row[4] if len(row) > 4 else "")
        system_qty = parse_int(row[5] if len(row) > 5 else 0)

        if not product_code:
            continue

        system_grade_counts = {"S+S": 0, "SS": 0, "SA": 0, "AS": 0, "AA": 0, "B": 0, "등급미정": 0}
        if "B급" in category:
            system_grade_counts["B"] = system_qty
        else:
            system_grade_counts["등급미정"] = system_qty

        products.append(
            {
                "id": f"excel-{index:04d}",
                "maker": infer_maker(product_code, description),
                "productCode": product_code,
                "name": product_code,
                "description": description,
                "category": category,
                "price": price,
                "exposure": exposure,
                "grade": "",
                "systemGradeCounts": system_grade_counts,
                "gradeCounts": {"S+S": 0, "SS": 0, "SA": 0, "AS": 0, "AA": 0, "B": 0},
                "systemQty": system_qty,
                "countQty": 0,
                "qty": 0,
                "note": "",
                "createdBy": "excel-import",
                "updatedBy": "excel-import",
                "checked": False,
                "checkedBy": "",
                "checkedAt": "",
                "updatedAt": now,
            }
        )

    state = load_existing_state()
    state["products"] = products
    state["movements"] = []
    state["stocktake"] = {"counts": {}, "updatedAt": now}
    return state


def main() -> int:
    if len(sys.argv) != 2:
        print("Usage: python tools/import_inventory_excel.py <excel-file>")
        return 2

    excel_path = Path(sys.argv[1]).resolve()
    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}")
        return 1

    if DATA_FILE.exists():
        backup = DATA_FILE.with_name(f"data.backup.{datetime.now().strftime('%Y%m%d-%H%M%S')}.json")
        shutil.copy2(DATA_FILE, backup)
        print(f"backup: {backup}")

    state = import_excel(excel_path)
    DATA_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"imported: {len(state['products'])} products")
    print(f"target: {DATA_FILE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
